# Make sure you close the excel file before you run this, won't work otherwise

# Using the openxl library we downloaded
# Importing the library, setting its new name to xl so we don't have to type it all out
import openpyxl as xl
# Importing the Graphs library from openpyxl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)  # Getting the Excel file
    # Getting the excel sheet
    sheet = wb['Sheet1']

    # Getting the first cell
    # They both do the same thing
    # cell = sheet['a1']
    # cell = sheet.cell(1, 1)
    # print(cell.value)  # Printing the cells value
    # print(sheet.max_row)  # Finding out how many rows are in the sheet

    # Iterating through the rows of the sheet
    # Staring from 2 because we don't want the headings
    for row in range(2, sheet.max_row + 1):
        # For each row, get the column at index 3
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9  # Getting 90% of the cells value
        # Get a new cell as the row position, in the 4th columns
        corrected_price_cell = sheet.cell(row, 4)
        # Set the value of the new cell we created to its correct value
        corrected_price_cell.value = corrected_price

    # For drawing the Bar chart

    # We use the reference class to select a range of values
    # We want to select the cell from row 2 to 4, in column 4
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    chart = BarChart()  # Creating an instance of the chart
    chart.add_data(values)  # Adding our data to the chart
    # Adding the chart to our sheet, on the e2 column
    sheet.add_chart(chart, 'e2')  # You might see 2 charts because i made another one in f2

    # Saving the file
    # You can see the file in your directories on the left
    wb.save(filename)


# Calling the function with the filename
process_workbook('transactions.xlsx')
